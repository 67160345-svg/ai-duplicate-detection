"""Evaluate precomputed duplicate-detection signals from CSV or Excel datasets."""

import argparse
import csv
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, Iterable, List

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from modules.scoring_engine import (
    DEFAULT_SIMILARITY_THRESHOLDS,
    classify_dataset_record,
)


SPLITS = {
    "03_Training_Dataset": "training",
    "04_Validation_Dataset": "validation",
    "05_Testing_Dataset": "testing",
}


def _find_header(rows: Iterable[List[object]]) -> tuple[int, List[str]]:
    for row_number, row in enumerate(rows):
        headers = [str(value).strip() if value is not None else "" for value in row]
        if "Ground_Truth" in headers and (
            "Embedding_Similarity" in headers or "Similarity_Score" in headers
        ):
            return row_number, headers
    raise ValueError("Could not find a dataset header row")


def _read_csv(path: Path) -> Dict[str, List[Dict[str, object]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        rows = list(csv.reader(source))
    header_number, headers = _find_header(rows)
    return {path.stem: _rows_to_records(headers, rows[header_number + 1 :])}


def _read_excel(path: Path) -> Dict[str, List[Dict[str, object]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Install openpyxl to evaluate Excel workbooks") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    datasets = {}
    for sheet_name, split_name in SPLITS.items():
        if sheet_name not in workbook.sheetnames:
            continue
        rows = workbook[sheet_name].iter_rows(values_only=True)
        buffered_rows = list(rows)
        header_number, headers = _find_header(buffered_rows)
        datasets[split_name] = _rows_to_records(
            headers, buffered_rows[header_number + 1 :]
        )
    if not datasets:
        raise ValueError("Workbook has no supported training/validation/testing sheets")
    return datasets


def _rows_to_records(headers: List[str], rows: Iterable[Iterable[object]]) -> List[Dict[str, object]]:
    records = []
    for row in rows:
        values = list(row)
        if not any(value is not None and str(value).strip() for value in values):
            continue
        records.append({header: values[index] if index < len(values) else None for index, header in enumerate(headers) if header})
    return records


def _metrics(
    records: List[Dict[str, object]], thresholds: Dict[str, float]
) -> Dict[str, object]:
    labels = sorted({str(record.get("Ground_Truth", "")).upper() for record in records if record.get("Ground_Truth")})
    confusion = {label: Counter() for label in labels}
    by_scenario = defaultdict(list)
    for record in records:
        truth = str(record.get("Ground_Truth", "")).upper()
        predicted = classify_dataset_record(record, thresholds)
        confusion.setdefault(truth, Counter())[predicted] += 1
        by_scenario[str(record.get("Scenario_Type", "UNKNOWN"))].append((truth, predicted))

    total = sum(sum(row.values()) for row in confusion.values())
    correct = sum(row.get(label, 0) for label, row in confusion.items())
    per_class = {}
    for label in labels:
        true_positive = confusion[label].get(label, 0)
        predicted_total = sum(row.get(label, 0) for row in confusion.values())
        actual_total = sum(confusion[label].values())
        precision = true_positive / predicted_total if predicted_total else 0.0
        recall = true_positive / actual_total if actual_total else 0.0
        f1 = 2 * precision * recall / (precision + recall) if precision + recall else 0.0
        per_class[label] = {
            "support": actual_total,
            "precision": round(precision, 4),
            "recall": round(recall, 4),
            "f1": round(f1, 4),
        }

    scenario_metrics = {}
    for scenario, pairs in sorted(by_scenario.items()):
        scenario_metrics[scenario] = {
            "count": len(pairs),
            "accuracy": round(sum(truth == predicted for truth, predicted in pairs) / len(pairs), 4),
        }
    macro_f1 = sum(item["f1"] for item in per_class.values()) / len(per_class) if per_class else 0.0
    return {
        "thresholds": thresholds,
        "records": total,
        "accuracy": round(correct / total, 4) if total else 0.0,
        "macro_f1": round(macro_f1, 4),
        "per_class": per_class,
        "by_scenario": scenario_metrics,
        "confusion_matrix": {label: dict(row) for label, row in confusion.items()},
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("dataset", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()

    datasets = _read_excel(args.dataset) if args.dataset.suffix.lower() in {".xlsx", ".xlsm"} else _read_csv(args.dataset)
    legacy_thresholds = {
        "exact_duplicate": 0.99,
        "near_duplicate": 0.85,
        "possible_duplicate": 0.85,
        "review": 0.70,
    }
    report = {
        split: {
            "legacy": _metrics(records, legacy_thresholds),
            "phase1": _metrics(records, DEFAULT_SIMILARITY_THRESHOLDS),
        }
        for split, records in datasets.items()
    }
    encoded = json.dumps(report, ensure_ascii=False, indent=2)
    print(encoded)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(encoded + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
